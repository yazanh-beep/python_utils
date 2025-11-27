import pandas as pd
import json
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# ============================================================================
# CONFIGURATION - Edit these variables as needed
# ============================================================================

# Input files
CAMERA_INVENTORY_JSON = 'camera_inventory_781cameras_2aggs_20251016_114305.json'
DIRECTORY_REPORT_EXCEL = 'directory_report.xlsx'
TRACKER_EXCEL = 'camera-switch-tracker.xlsx'

# Sheet names
DIRECTORY_SHEET_NAME = 'Sheet1'
TRACKER_CAMERA_SHEET = 'camera'

# Output file
OUTPUT_EXCEL = 'switch_camera_tracker.xlsx'

# Column names in directory report (adjust if your Excel has different column names)
DIR_COL_CAMERA_NAME = 'Camera stream name'
DIR_COL_CAMERA_STATE = 'Camera state/status'
DIR_COL_IP_ADDRESS = 'IP address'
DIR_COL_MAC_ADDRESS = 'MAC address'

# ============================================================================

def normalize_mac(mac_address):
    """Normalize MAC address format to uppercase with colons"""
    if pd.isna(mac_address) or mac_address == '':
        return None
    # Remove any separators and convert to uppercase
    mac = str(mac_address).upper().replace(':', '').replace('-', '').replace('.', '')
    # Add colons back in standard format
    return ':'.join([mac[i:i+2] for i in range(0, len(mac), 2)])

# Define highlight colors
YELLOW_FILL = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')  # No name found
ORANGE_FILL = PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid')  # No switch info found
LIGHTBLUE_FILL = PatternFill(start_color='ADD8E6', end_color='ADD8E6', fill_type='solid')  # Duplicate MAC (same camera, different names)
RED_FILL = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')  # Duplicate IP with different MAC

def main():
    # Read the JSON file with camera inventory (switch info)
    print(f"Reading {CAMERA_INVENTORY_JSON}...")
    with open(CAMERA_INVENTORY_JSON, 'r') as f:
        camera_inventory_data = json.load(f)
    
    # Debug: Print the type and keys to understand the structure
    print(f"Data type: {type(camera_inventory_data)}")
    if isinstance(camera_inventory_data, dict):
        print(f"Keys in JSON: {list(camera_inventory_data.keys())}")
    
    # Handle both old format (list) and new format (dict with metadata)
    if isinstance(camera_inventory_data, dict):
        if 'cameras' in camera_inventory_data:
            # New format with metadata
            camera_inventory = camera_inventory_data['cameras']
            metadata = camera_inventory_data.get('discovery_metadata', {})
            
            print("\n" + "="*60)
            print("CAMERA DISCOVERY METADATA")
            print("="*60)
            print(f"Discovery timestamp: {metadata.get('timestamp', 'N/A')}")
            print(f"Seed switch: {metadata.get('seed_switch', 'N/A')}")
            print(f"Total cameras found: {metadata.get('total_cameras', len(camera_inventory))}")
            print(f"Total aggregates scanned: {metadata.get('total_aggregates', 'N/A')}")
            
            if 'aggregates' in metadata:
                print(f"\nAggregates discovered:")
                for agg in metadata['aggregates']:
                    seed_marker = " (SEED)" if agg.get('is_seed') else ""
                    print(f"  - {agg.get('hostname', 'Unknown'):<40} {agg.get('ip')}{seed_marker}")
            print("="*60 + "\n")
        else:
            # Dict but no 'cameras' key - might be old format stored as dict
            # Check if it has expected fields
            first_key = list(camera_inventory_data.keys())[0] if camera_inventory_data else None
            if first_key and isinstance(camera_inventory_data[first_key], dict):
                # It's a dict of cameras (unusual format)
                camera_inventory = list(camera_inventory_data.values())
            else:
                print("ERROR: Unexpected JSON structure")
                print(f"Please check the format of {CAMERA_INVENTORY_JSON}")
                return
    elif isinstance(camera_inventory_data, list):
        # Old format - just a list
        camera_inventory = camera_inventory_data
    else:
        print(f"ERROR: Unexpected data type in JSON: {type(camera_inventory_data)}")
        return
    
    # Create a dictionary with MAC address as key
    inventory_dict = {}
    switch_type_stats = {}
    
    for entry in camera_inventory:
        mac = normalize_mac(entry['mac_address'])
        if mac:
            switch_type = entry.get('switch_type', 'UNKNOWN')
            inventory_dict[mac] = {
                'switch_name': entry['switch_name'],
                'switch_type': switch_type,
                'port': entry['port'],
                'vlan': entry.get('vlan', '')
            }
            # Track statistics by switch type
            switch_type_stats[switch_type] = switch_type_stats.get(switch_type, 0) + 1
    
    print(f"Loaded {len(inventory_dict)} camera records from inventory")
    
    if switch_type_stats:
        print("\nCameras by switch type:")
        for switch_type, count in sorted(switch_type_stats.items()):
            print(f"  - {switch_type}: {count}")
    
    # Read the directory report Excel file
    print(f"\nReading {DIRECTORY_REPORT_EXCEL}...")
    directory_df = pd.read_excel(DIRECTORY_REPORT_EXCEL, sheet_name=DIRECTORY_SHEET_NAME)
    
    print(f"Total rows in directory report: {len(directory_df)}")
    
    # IMPORTANT: We only filter out formatting/status rows, NOT duplicate cameras
    # Duplicate cameras (same MAC/IP, different names) are kept intentionally
    original_count = len(directory_df)
    
    # Remove rows where camera name is NaN or empty (these are formatting rows)
    directory_df = directory_df[directory_df[DIR_COL_CAMERA_NAME].notna()]
    directory_df = directory_df[directory_df[DIR_COL_CAMERA_NAME].astype(str).str.strip() != '']
    
    # Remove rows that are ONLY status updates (empty camera name + "Streaming" status)
    # These are formatting rows, not actual duplicate cameras
    directory_df = directory_df[
        ~((directory_df[DIR_COL_CAMERA_STATE] == 'Streaming') & 
          (directory_df[DIR_COL_CAMERA_NAME].astype(str).str.strip() == ''))
    ]
    
    filtered_count = len(directory_df)
    removed_count = original_count - filtered_count
    
    print(f"Filtered out {removed_count} empty/status formatting rows")
    print(f"Processing {filtered_count} camera records (including duplicates)")
    
    # Check for duplicate MACs to inform user
    directory_df['MAC_normalized'] = directory_df[DIR_COL_MAC_ADDRESS].apply(normalize_mac)
    duplicate_macs = directory_df[directory_df['MAC_normalized'].notna()].duplicated(subset=['MAC_normalized'], keep=False).sum()
    if duplicate_macs > 0:
        print(f"Note: Found {duplicate_macs} rows with duplicate MAC addresses (will be highlighted in light blue)")
    
    # Check for duplicate IPs with different MACs
    ip_groups = directory_df[directory_df[DIR_COL_IP_ADDRESS].notna()].groupby(DIR_COL_IP_ADDRESS)['MAC_normalized'].nunique()
    duplicate_ips = (ip_groups > 1).sum()
    if duplicate_ips > 0:
        print(f"WARNING: Found {duplicate_ips} IP addresses shared by different MAC addresses (will be highlighted in RED)")
    
    # Load the tracker workbook
    print(f"\nLoading {TRACKER_EXCEL}...")
    wb = load_workbook(TRACKER_EXCEL)
    ws = wb[TRACKER_CAMERA_SHEET]
    
    # Clear existing data (except header)
    print("Clearing existing data in camera sheet...")
    ws.delete_rows(2, ws.max_row)
    
    # Prepare data to write
    print("\nMatching cameras and preparing data...")
    row_num = 2
    matched_count = 0
    no_switch_info_count = 0
    no_name_info_count = 0
    duplicate_mac_count = 0
    duplicate_ip_count = 0
    
    # Track which MACs from inventory we've used
    used_macs = set()
    
    # Track MACs we've already written to detect duplicates
    written_macs = {}  # mac -> list of (row_num, camera_name, ip_address)
    
    # Track IPs we've already written to detect IP duplicates with different MACs
    written_ips = {}  # ip -> list of (row_num, mac_address)
    
    # First pass: Process all cameras from directory report
    for idx, cam_row in directory_df.iterrows():
        mac = cam_row['MAC_normalized']
        camera_name = cam_row[DIR_COL_CAMERA_NAME]
        ip_address = cam_row[DIR_COL_IP_ADDRESS]
        original_mac = cam_row[DIR_COL_MAC_ADDRESS]
        
        # Check if this MAC was already written (duplicate detection)
        is_duplicate_mac = mac in written_macs if mac else False
        
        # Check if this MAC has a DIFFERENT IP than previous entries
        is_same_mac_diff_ip = False
        if is_duplicate_mac and ip_address:
            for prev_row, prev_name, prev_ip in written_macs[mac]:
                if prev_ip and prev_ip != ip_address:
                    is_same_mac_diff_ip = True
                    break
        
        # Check if this IP was already written with a DIFFERENT MAC
        is_duplicate_ip_diff_mac = False
        if ip_address and ip_address in written_ips:
            for prev_row, prev_mac in written_ips[ip_address]:
                if prev_mac != mac:
                    is_duplicate_ip_diff_mac = True
                    break
        
        # Look up switch info from inventory
        switch_info = inventory_dict.get(mac)
        
        if switch_info:
            # Write to Excel - fully matched
            ws.cell(row=row_num, column=1, value=camera_name)
            ws.cell(row=row_num, column=2, value=original_mac)
            ws.cell(row=row_num, column=3, value=ip_address)
            
            # Enhanced: Include switch type in the switch name display
            switch_type = switch_info.get('switch_type', 'UNKNOWN')
            switch_display = f"{switch_info['switch_name']} [{switch_type}]"
            ws.cell(row=row_num, column=4, value=switch_display)
            ws.cell(row=row_num, column=5, value=switch_info['port'])
            
            # Priority order for highlighting: RED > LIGHT BLUE
            if is_duplicate_ip_diff_mac or is_same_mac_diff_ip:
                # Network conflict - RED (most critical)
                for col in range(1, 6):
                    ws.cell(row=row_num, column=col).fill = RED_FILL
                duplicate_ip_count += 1
                
                if is_duplicate_ip_diff_mac:
                    for prev_row, prev_mac in written_ips[ip_address]:
                        if prev_mac != mac:
                            for col in range(1, 6):
                                ws.cell(row=prev_row, column=col).fill = RED_FILL
                    
                    print(f"  RED: Duplicate IP {ip_address} with different MAC:")
                    print(f"    - Current MAC: {original_mac}")
                    print(f"    - Camera: {camera_name}")
                
                if is_same_mac_diff_ip:
                    for prev_row, prev_name, prev_ip in written_macs[mac]:
                        if prev_ip and prev_ip != ip_address:
                            for col in range(1, 6):
                                ws.cell(row=prev_row, column=col).fill = RED_FILL
                    
                    print(f"  RED: Same MAC {original_mac} with different IP:")
                    print(f"    - Current IP: {ip_address}")
                    print(f"    - Camera: {camera_name}")
                
            elif is_duplicate_mac:
                # Duplicate MAC with same IP (same camera, different names) - LIGHT BLUE
                for col in range(1, 6):
                    ws.cell(row=row_num, column=col).fill = LIGHTBLUE_FILL
                duplicate_mac_count += 1
                prev_entries = written_macs[mac]
                print(f"  LIGHT BLUE: Duplicate MAC {original_mac} found:")
                print(f"    - Previous: {prev_entries[-1][1]}")
                print(f"    - Current:  {camera_name}")
                # Also highlight the previous entries
                for prev_row, prev_name, prev_ip in prev_entries:
                    for col in range(1, 6):
                        ws.cell(row=prev_row, column=col).fill = LIGHTBLUE_FILL
            
            matched_count += 1
            if mac:
                used_macs.add(mac)
                if mac not in written_macs:
                    written_macs[mac] = []
                written_macs[mac].append((row_num, camera_name, ip_address))
            
            # Track IPs
            if ip_address:
                if ip_address not in written_ips:
                    written_ips[ip_address] = []
                written_ips[ip_address].append((row_num, mac))
                
        else:
            # Write partial data - have name but no switch info (highlight ORANGE)
            ws.cell(row=row_num, column=1, value=camera_name)
            ws.cell(row=row_num, column=2, value=original_mac)
            ws.cell(row=row_num, column=3, value=ip_address)
            ws.cell(row=row_num, column=4, value='NOT FOUND')
            ws.cell(row=row_num, column=5, value='NOT FOUND')
            
            # Check for network conflicts even without switch info
            if is_duplicate_ip_diff_mac or is_same_mac_diff_ip:
                # RED takes priority over ORANGE
                for col in range(1, 6):
                    ws.cell(row=row_num, column=col).fill = RED_FILL
                duplicate_ip_count += 1
                
                if is_duplicate_ip_diff_mac:
                    for prev_row, prev_mac in written_ips[ip_address]:
                        if prev_mac != mac:
                            for col in range(1, 6):
                                ws.cell(row=prev_row, column=col).fill = RED_FILL
                    
                    print(f"  RED: Duplicate IP {ip_address} with different MAC (no switch info):")
                    print(f"    - Current MAC: {original_mac}")
                
                if is_same_mac_diff_ip:
                    for prev_row, prev_name, prev_ip in written_macs[mac]:
                        if prev_ip and prev_ip != ip_address:
                            for col in range(1, 6):
                                ws.cell(row=prev_row, column=col).fill = RED_FILL
                    
                    print(f"  RED: Same MAC {original_mac} with different IP (no switch info):")
                    print(f"    - Current IP: {ip_address}")
            else:
                # Highlight the entire row in ORANGE
                for col in range(1, 6):
                    ws.cell(row=row_num, column=col).fill = ORANGE_FILL
            
            no_switch_info_count += 1
            print(f"  ORANGE: No switch info found for {camera_name} (MAC: {original_mac})")
            
            # Track written MACs and IPs even if no switch info found
            if mac:
                if mac not in written_macs:
                    written_macs[mac] = []
                written_macs[mac].append((row_num, camera_name, ip_address))
            
            if ip_address:
                if ip_address not in written_ips:
                    written_ips[ip_address] = []
                written_ips[ip_address].append((row_num, mac))
        
        row_num += 1
    
    # Second pass: Add cameras from inventory that don't have names (not in directory report)
    print("\nChecking for cameras in inventory without names...")
    for mac, switch_info in inventory_dict.items():
        if mac not in used_macs:
            # This MAC has switch info but no camera name (highlight YELLOW)
            ws.cell(row=row_num, column=1, value='NAME NOT FOUND')
            ws.cell(row=row_num, column=2, value=mac)
            ws.cell(row=row_num, column=3, value='')
            
            # Include switch type in display
            switch_type = switch_info.get('switch_type', 'UNKNOWN')
            switch_display = f"{switch_info['switch_name']} [{switch_type}]"
            ws.cell(row=row_num, column=4, value=switch_display)
            ws.cell(row=row_num, column=5, value=switch_info['port'])
            
            # Highlight the entire row in YELLOW
            for col in range(1, 6):
                ws.cell(row=row_num, column=col).fill = YELLOW_FILL
            
            no_name_info_count += 1
            print(f"  YELLOW: No camera name found for MAC: {mac} on {switch_info['switch_name']} [{switch_type}] port {switch_info['port']}")
            
            row_num += 1
    
    # Save the workbook
    print(f"\nSaving updated tracker to {OUTPUT_EXCEL}...")
    wb.save(OUTPUT_EXCEL)
    
    # Print summary
    print("\n" + "="*60)
    print("SUMMARY")
    print("="*60)
    print(f"Total cameras from directory: {len(directory_df)}")
    print(f"Total MACs from inventory: {len(inventory_dict)}")
    print(f"\n✓ Successfully matched: {matched_count}")
    print(f"⚠  ORANGE - Have name but no switch info: {no_switch_info_count}")
    print(f"⚠  YELLOW - Have switch info but no name: {no_name_info_count}")
    print(f"⚠  LIGHT BLUE - Duplicate MAC (same camera, different names): {duplicate_mac_count}")
    print(f"⚠  RED - Duplicate IP with different MACs (IP CONFLICT): {duplicate_ip_count}")
    print(f"\nTotal rows written: {row_num - 2}")
    print(f"\nOutput saved to: {OUTPUT_EXCEL}")
    print("\nColor coding:")
    print("  - No color: Fully matched (name + switch info)")
    print("  - ORANGE: Camera name found but no switch info")
    print("  - YELLOW: Switch info found but no camera name")
    print("  - LIGHT BLUE: Same MAC/IP with different names (needs reconciliation)")
    print("  - RED: Network conflict (CRITICAL!)")
    print("      * Same IP with different MACs, OR")
    print("      * Same MAC with different IPs")
    print("\nNOTE: Switch names now include switch type in brackets [EDGE/SERVER/OTHER]")
    print("      Red entries indicate network configuration errors that need immediate attention!")
    print("="*60)

if __name__ == "__main__":
    main()
