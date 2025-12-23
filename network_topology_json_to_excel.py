import json
import openpyxl
from collections import defaultdict

def load_json_data(json_file):
    """Load the network topology JSON file."""
    with open(json_file, 'r') as f:
        return json.load(f)

def clean_hostname(hostname):
    """
    Clean hostname by removing .CAM.INT or other domain suffixes if present.
    Returns the base hostname.
    """
    if not hostname:
        return hostname
    # Remove common domain suffixes
    for suffix in ['.CAM.INT', '.cam.int', '.local', '.cisco.com']:
        if hostname.endswith(suffix):
            return hostname[:-len(suffix)]
    return hostname

def is_valid_uplink(neighbor):
    """
    Determines if a neighbor is a valid uplink/aggregate switch.
    
    Criteria:
    1. Must have a Management IP (filters out dumb endpoints/cameras).
    2. Must not be the same device (loopback check).
    """
    # In your data, cameras/endpoints have null management IPs. 
    # Real switches have IPs.
    if neighbor.get('neighbor_mgmt_ip'):
        return True
    return False

def format_uplinks(neighbors, current_hostname):
    """
    Format switch uplinks. 
    scans all neighbors and identifies any valid switch as an uplink.
    """
    uplink_connections = defaultdict(list)
    clean_current = clean_hostname(current_hostname)
    
    for neighbor in neighbors:
        neighbor_hostname = neighbor.get('neighbor_hostname', '')
        
        # LOGIC CHANGE: We simply check if it's a valid switch neighbor
        if is_valid_uplink(neighbor):
            clean_name = clean_hostname(neighbor_hostname)
            
            # Avoid listing ourselves if the data contains loopbacks
            if clean_name == clean_current:
                continue
            
            connection_info = {
                'local_port': neighbor.get('local_interface', 'Unknown'),
                'remote_port': neighbor.get('remote_interface', 'Unknown'),
                'agg_full_name': clean_name
            }
            uplink_connections[clean_name].append(connection_info)
    
    if not uplink_connections:
        return "", ""
    
    # Format the output
    agg_names = []
    uplink_details = []
    
    for agg_hostname, connections in sorted(uplink_connections.items()):
        agg_names.append(agg_hostname)
        
        if len(connections) == 1:
            # Single connection
            conn = connections[0]
            uplink_details.append(f"On {agg_hostname}: Local {conn['local_port']} -> Remote {conn['remote_port']}")
        else:
            # Multiple connections to same switch (Port Channel / Redundant links)
            ports_info = []
            for idx, conn in enumerate(connections, 1):
                ports_info.append(f"Link {idx}: Local {conn['local_port']} -> Remote {conn['remote_port']}")
            uplink_details.append(f"On {agg_hostname}: {'; '.join(ports_info)}")
    
    aggregate_switch_str = " and ".join(agg_names)
    uplink_port_str = " | ".join(uplink_details)
    
    return aggregate_switch_str, uplink_port_str

def populate_excel_tracker(json_file, excel_file, output_file):
    """
    Populate the Excel tracker with data from the JSON file.
    """
    # Load data
    topology_data = load_json_data(json_file)
    
    # Load Excel workbook
    try:
        wb = openpyxl.load_workbook(excel_file)
    except FileNotFoundError:
        print(f"Error: Could not find template file: {excel_file}")
        return

    # Access the 'switch' sheet
    if 'switch' not in wb.sheetnames:
        print("Error: 'switch' sheet not found in workbook")
        return
    
    ws = wb['switch']
    
    # Start writing from row 2 (row 1 has headers)
    current_row = 2
    
    for device in topology_data:
        hostname = device.get('hostname', '')
        serial_number = device.get('serial_number', '')
        mgmt_ip = device.get('management_ip', '')
        switch_model = device.get('switch_model', '')
        ios_version = device.get('ios_version', '')
        neighbors = device.get('neighbors', [])
        
        # Unified logic: Just find the uplink(s)
        aggregate_switch, uplink_port = format_uplinks(neighbors, hostname)
        
        # Write to Excel
        ws.cell(row=current_row, column=1, value=hostname)
        ws.cell(row=current_row, column=2, value=serial_number)
        ws.cell(row=current_row, column=3, value=mgmt_ip)
        ws.cell(row=current_row, column=4, value=switch_model)
        ws.cell(row=current_row, column=5, value=ios_version)
        ws.cell(row=current_row, column=6, value=aggregate_switch)
        ws.cell(row=current_row, column=7, value=uplink_port)
        
        current_row += 1
    
    # Save the workbook
    wb.save(output_file)
    print(f"Successfully populated {current_row - 2} switches in {output_file}")

if __name__ == "__main__":
    # File paths
    json_file = "network_topology.json"
    excel_file = "camera-switch-tracker-template.xlsx"
    output_file = "camera-switch-tracker.xlsx"
    
    # Populate the tracker
    populate_excel_tracker(json_file, excel_file, output_file)
