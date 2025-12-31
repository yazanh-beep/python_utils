import json
import openpyxl
from collections import defaultdict

def load_json_data(json_file):
    """Load the network topology JSON file."""
    with open(json_file, 'r') as f:
        return json.load(f)

def clean_hostname(hostname):
    """
    Clean hostname by removing .CAM.INT or other domain suffixes if present.
    Returns the base hostname.
    """
    if not hostname:
        return ""
    # Remove common domain suffixes
    for suffix in ['.CAM.INT', '.cam.int', '.local', '.cisco.com', '.simplex.net', '.jci.net', '.JCI.net']:
        if hostname.endswith(suffix):
            return hostname[:-len(suffix)]
    return hostname

def is_valid_uplink(neighbor):
    """
    Determines if a neighbor is a valid uplink/aggregate switch.
    """
    # If it has an IP, we treat it as a valid switch neighbor
    if neighbor.get('neighbor_mgmt_ip'):
        return True
    
    # Also treat it as valid if we are inferring relationships for unvisited devices
    # (The calling logic determines context)
    return False

def format_uplinks(neighbors, current_hostname):
    """
    Format switch uplinks for a main device entry.
    """
    uplink_connections = defaultdict(list)
    clean_current = clean_hostname(current_hostname)
    
    for neighbor in neighbors:
        if is_valid_uplink(neighbor):
            neighbor_hostname = neighbor.get('neighbor_hostname', '')
            clean_name = clean_hostname(neighbor_hostname)
            
            if clean_name == clean_current:
                continue
            
            connection_info = {
                'local_port': neighbor.get('local_interface', 'Unknown'),
                'remote_port': neighbor.get('remote_interface', 'Unknown'),
                'agg_full_name': clean_name
            }
            uplink_connections[clean_name].append(connection_info)
    
    return generate_uplink_strings(uplink_connections)

def generate_uplink_strings(uplink_connections):
    """Helper to convert connection dict to string format."""
    if not uplink_connections:
        return "", ""
    
    agg_names = []
    uplink_details = []
    
    for agg_hostname, connections in sorted(uplink_connections.items()):
        agg_names.append(agg_hostname)
        
        if len(connections) == 1:
            conn = connections[0]
            # Standard format: On [Switch]: Local [Port] -> Remote [Port]
            uplink_details.append(f"On {agg_hostname}: Local {conn['local_port']} -> Remote {conn['remote_port']}")
        else:
            ports_info = []
            for idx, conn in enumerate(connections, 1):
                ports_info.append(f"Link {idx}: Local {conn['local_port']} -> Remote {conn['remote_port']}")
            uplink_details.append(f"On {agg_hostname}: {'; '.join(ports_info)}")
    
    return " and ".join(agg_names), " | ".join(uplink_details)

def find_unvisited_neighbors(topology_data):
    """
    Scans all neighbors to find devices that were NOT visited (no main entry).
    Returns a dict of {hostname: uplink_data} for these unvisited devices.
    """
    visited_hostnames = set()
    for device in topology_data:
        h = clean_hostname(device.get('hostname'))
        if h: visited_hostnames.add(h)

    unvisited = {}

    for device in topology_data:
        parent_hostname = clean_hostname(device.get('hostname'))
        
        for neighbor in device.get('neighbors', []):
            nbr_raw_name = neighbor.get('neighbor_hostname')
            nbr_clean_name = clean_hostname(nbr_raw_name)
            
            # If this neighbor is NOT in our main list of visited devices
            if nbr_clean_name and nbr_clean_name not in visited_hostnames:
                
                # We inferred the uplink! The uplink is the device we are currently looking at.
                # BUT we must swap ports: 
                # The "Local" port in the JSON is the parent's port. 
                # The "Remote" port in the JSON is this unvisited device's port.
                
                if nbr_clean_name not in unvisited:
                    unvisited[nbr_clean_name] = defaultdict(list)
                
                connection_info = {
                    # From the perspective of the unvisited switch:
                    # Its "Local" is the parent's "Remote"
                    # Its "Remote" is the parent's "Local"
                    'local_port': neighbor.get('remote_interface', 'Unknown'), 
                    'remote_port': neighbor.get('local_interface', 'Unknown'),
                    'agg_full_name': parent_hostname
                }
                unvisited[nbr_clean_name][parent_hostname].append(connection_info)

    return unvisited

def populate_excel_tracker(json_file, excel_file, output_file):
    """
    Populate the Excel tracker with data from the JSON file.
    """
    topology_data = load_json_data(json_file)
    
    try:
        wb = openpyxl.load_workbook(excel_file)
    except FileNotFoundError:
        print(f"Error: Could not find template file: {excel_file}")
        return

    if 'switch' not in wb.sheetnames:
        print("Error: 'switch' sheet not found in workbook")
        return
    
    ws = wb['switch']
    current_row = 2
    
    # 1. Process Main (Visited) Devices
    for device in topology_data:
        hostname = clean_hostname(device.get('hostname', ''))
        
        aggregate_switch, uplink_port = format_uplinks(device.get('neighbors', []), hostname)
        
        ws.cell(row=current_row, column=1, value=hostname)
        ws.cell(row=current_row, column=2, value=device.get('serial_number', ''))
        ws.cell(row=current_row, column=3, value=device.get('management_ip', ''))
        ws.cell(row=current_row, column=4, value=device.get('switch_model', ''))
        ws.cell(row=current_row, column=5, value=device.get('ios_version', ''))
        ws.cell(row=current_row, column=6, value=aggregate_switch)
        ws.cell(row=current_row, column=7, value=uplink_port)
        
        current_row += 1

    # 2. Process Unvisited Neighbors (e.g., "SWITCH", "IDF_1" without IP)
    unvisited_data = find_unvisited_neighbors(topology_data)
    
    print(f"Found {len(unvisited_data)} unvisited/unmanaged switches to add...")

    for hostname, uplink_dict in sorted(unvisited_data.items()):
        # Generate the uplink string using our helper
        agg_str, port_str = generate_uplink_strings(uplink_dict)
        
        ws.cell(row=current_row, column=1, value=hostname)
        ws.cell(row=current_row, column=2, value="N/A (Unmanaged/No IP)")
        ws.cell(row=current_row, column=3, value="Unknown") # IP Unknown
        ws.cell(row=current_row, column=4, value="Unknown") # Model Unknown
        ws.cell(row=current_row, column=5, value="Unknown") # Version Unknown
        ws.cell(row=current_row, column=6, value=agg_str)
        ws.cell(row=current_row, column=7, value=port_str)
        
        current_row += 1
    
    wb.save(output_file)
    print(f"Successfully populated {current_row - 2} switches in {output_file}")

if __name__ == "__main__":
    json_file = "network_topology.json"
    excel_file = "camera-switch-tracker-template.xlsx"
    output_file = "camera-switch-tracker.xlsx"
    
    populate_excel_tracker(json_file, excel_file, output_file)
