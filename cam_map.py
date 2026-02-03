import pandas as pd
import json
import glob
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# ============================================================================
# CONFIGURATION
# ============================================================================
DIRECTORY_REPORT_EXCEL = 'directory_report.xlsx'
TRACKER_EXCEL = 'camera-switch-tracker.xlsx'
OUTPUT_EXCEL = 'camera-switch-tracker.xlsx'

# Keywords to find the correct Header Row
# We look for a row containing "MAC Address" to know where data starts
HEADER_ANCHOR = "MAC address"

# Target Columns in Output (Do not change)
HEADERS_OUTPUT = [
    "Camera Name",      # Col 1
    "MAC Address",      # Col 2
    "IP Address",       # Col 3
    "Switch Name",      # Col 4
    "Port",             # Col 5
    "Location",         # Col 6
    "Cloud Cam Exporter", # Col 7
    "Status"            # Col 8
]

# Colors
YELLOW_FILL = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
ORANGE_FILL = PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid')

def normalize_mac(mac_address):
    if pd.isna(mac_address) or str(mac_address).strip() == '':
        return None
    mac = str(mac_address).upper().replace(':', '').replace('-', '').replace('.', '')
    # Check for valid hex characters to avoid junk data
    try:
        if len(mac) == 12:
            return ':'.join([mac[i:i+2] for i in range(0, len(mac), 2)])
    except: pass
    return None

def get_inventory_file():
    files = glob.glob('camera_inventory*.json')
    return max(files, key=os.path.getmtime) if files else None

def find_header_row(file_path):
    """
    Scans the first 10 rows to find the row that contains 'MAC Address'.
    Returns the dataframe with the correct header.
    """
    # Read first 10 rows without header
    try:
        df_preview = pd.read_excel(file_path, header=None, nrows=10)
    except:
        df_preview = pd.read_csv(file_path, header=None, nrows=10)

    header_row_index = 0
    found = False
    
    # Search for the anchor keyword
    for idx, row in df_preview.iterrows():
        row_str = row.astype(str).str.lower().values
        if any(HEADER_ANCHOR.lower() in s for s in row_str):
            header_row_index = idx
            found = True
            break
    
    if found:
        print(f"✓ Found headers on Row {header_row_index + 1}")
        return pd.read_excel(file_path, header=header_row_index)
    else:
        print("⚠ Warning: Could not find 'MAC Address' header row. Assuming Row 1.")
        return pd.read_excel(file_path, header=0)

def main():
    print("🚀 Starting Strict Update...")
    
    # 1. LOAD INVENTORY
    inventory_file = get_inventory_file()
    if not inventory_file:
        print("❌ ERROR: No inventory JSON found.")
        return
    
    with open(inventory_file, 'r') as f:
        data = json.load(f)
        inventory_list = data.get('cameras', data) if isinstance(data, dict) else data

    inventory_dict = {}
    for entry in inventory_list:
        if 'mac_address' in entry:
            mac = normalize_mac(entry['mac_address'])
            if mac: inventory_dict[mac] = entry

    # 2. READ REPORT (Using Header Hunt)
    print(f"Reading {DIRECTORY_REPORT_EXCEL}...")
    df = find_header_row(DIRECTORY_REPORT_EXCEL)
    
    # Normalize column names for easy lookup
    df.columns = df.columns.astype(str).str.strip().str.lower()
    
    # Map Columns Explicitly
    col_map = {
        'mac': next((c for c in df.columns if 'mac' in c), None),
        'name': next((c for c in df.columns if 'camera stream' in c or 'name' in c), None),
        'ip': next((c for c in df.columns if 'ip address' in c), None),
        'loc': next((c for c in df.columns if 'location' in c), None),
        'exp': next((c for c in df.columns if 'exporter' in c), None),
        'status': next((c for c in df.columns if 'status' in c), None)
    }

    print("Column Mapping (Verify this matches your data):")
    print(json.dumps(col_map, indent=2))

    if not col_map['mac']:
        print("❌ CRITICAL: Could not identify MAC address column.")
        return

    # 3. CONSOLIDATE DATA
    # Forward fill MACs to handle "split row" blocks
    df[col_map['mac']] = df[col_map['mac']].ffill()
    
    consolidated = {}
    
    for _, row in df.iterrows():
        mac = normalize_mac(row.get(col_map['mac']))
        if not mac: continue

        # Extract values safely
        v_name = row.get(col_map['name']) if col_map['name'] else ''
        v_ip = row.get(col_map['ip']) if col_map['ip'] else ''
        v_loc = row.get(col_map['loc']) if col_map['loc'] else ''
        v_exp = row.get(col_map['exp']) if col_map['exp'] else ''
        v_stat = row.get(col_map['status']) if col_map['status'] else ''

        if mac not in consolidated:
            consolidated[mac] = {
                'name': v_name, 'ip': v_ip, 'loc': v_loc, 
                'exp': v_exp, 'status': v_stat
            }
        else:
            # Always grab the LAST status (bottom of the block)
            if pd.notna(v_stat): consolidated[mac]['status'] = v_stat
            # Grab Name/Exporter only if missing (top of the block usually has these)
            if not consolidated[mac]['name'] and pd.notna(v_name): consolidated[mac]['name'] = v_name
            if not consolidated[mac]['exp'] and pd.notna(v_exp): consolidated[mac]['exp'] = v_exp
            if not consolidated[mac]['loc'] and pd.notna(v_loc): consolidated[mac]['loc'] = v_loc

    # 4. WRITE TO EXCEL
    wb = load_workbook(TRACKER_EXCEL)
    
    for sheet_name in ['camera', 'Server']:
        # Create sheet if missing
        if sheet_name not in wb.sheetnames:
            wb.create_sheet(sheet_name)
        
        ws = wb[sheet_name]
        
        # FORCE REWRITE HEADERS (Fixes "Missing Header" issue)
        ws.delete_rows(1, ws.max_row + 1) # Clear everything
        for c, h in enumerate(HEADERS_OUTPUT, 1):
            ws.cell(row=1, column=c, value=h)
            
    ws_cam = wb['camera']
    ws_srv = wb['Server']
    
    # Start writing at Row 2
    rows = {'camera': 2, 'Server': 2}
    used_macs = set()

    for mac, data in consolidated.items():
        inv = inventory_dict.get(mac)
        
        # Server Detection
        is_server = False
        sw_name, sw_port = "NOT FOUND", "NOT FOUND"
        fill = ORANGE_FILL
        
        if inv:
            s_name = str(inv.get('switch_name', '')).upper()
            s_type = str(inv.get('switch_type', '')).upper()
            if s_type == 'SERVER' or 'SERVER' in s_name:
                is_server = True
            
            sw_name = f"{inv.get('switch_name')} [{inv.get('switch_type')}]"
            sw_port = inv.get('port')
            used_macs.add(mac)
            fill = None # No fill if found

        # Determine target
        t_sheet = 'Server' if is_server else 'camera'
        ws = ws_srv if is_server else ws_cam
        r = rows[t_sheet]
        
        # Explicit Column Writing (Fixes "Status under Exporter" issue)
        # We manually map variables to columns 1-8
        ws.cell(row=r, column=1, value=data['name'])       # Col 1: Name
        ws.cell(row=r, column=2, value=mac)                # Col 2: MAC
        ws.cell(row=r, column=3, value=data['ip'])         # Col 3: IP
        ws.cell(row=r, column=4, value=sw_name)            # Col 4: Switch
        ws.cell(row=r, column=5, value=sw_port)            # Col 5: Port
        ws.cell(row=r, column=6, value=data['loc'])        # Col 6: Location
        ws.cell(row=r, column=7, value=data['exp'])        # Col 7: Exporter
        ws.cell(row=r, column=8, value=data['status'])     # Col 8: Status

        if fill:
            for c in range(1, 9): ws.cell(row=r, column=c).fill = fill
        
        rows[t_sheet] += 1

    # Remaining Inventory
    for mac, inv in inventory_dict.items():
        if mac not in used_macs:
            s_name = str(inv.get('switch_name', '')).upper()
            s_type = str(inv.get('switch_type', '')).upper()
            is_server = (s_type == 'SERVER' or 'SERVER' in s_name)
            
            t_sheet = 'Server' if is_server else 'camera'
            ws = ws_srv if is_server else ws_cam
            r = rows[t_sheet]
            
            ws.cell(row=r, column=1, value="INVENTORY ONLY")
            ws.cell(row=r, column=2, value=mac)
            ws.cell(row=r, column=4, value=inv.get('switch_name'))
            ws.cell(row=r, column=5, value=inv.get('port'))
            
            for c in range(1, 9): ws.cell(row=r, column=c).fill = YELLOW_FILL
            rows[t_sheet] += 1

    # Final Save
    wb.save(OUTPUT_EXCEL)
    print("✅ Done. Headers forced. Columns aligned.")

if __name__ == "__main__":
    main()
