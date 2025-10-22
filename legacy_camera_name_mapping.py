import pandas as pd
import json
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# ============================================================================
# CONFIGURATION - Edit these variables as needed
# ============================================================================

# Input files
CAMERA_INVENTORY_JSON = 'camera_inventory_1006cameras_20251021_110632.json'
DIRECTORY_REPORT_EXCEL = 'sln4a_directory_report.xlsx'
TRACKER_EXCEL = 'camera-switch-tracker-template.xlsx'

# Sheet names
DIRECTORY_SHEET_NAME = 'Sheet1'
TRACKER_CAMERA_SHEET = 'camera'

# Output file
OUTPUT_EXCEL = 'switch_camera_tracker_UPDATED.xlsx'

# Column names in the new directory report format
DIR_COL_CAMERA_NAME = 'System Device Name\n(Location is required)'
DIR_COL_IP_ADDRESS = 'IP Address\n(x.x.x.x)'
DIR_COL_MAC_ADDRESS = 'MAC Address\n(xxxxxxxxxxxx)'
DIR_COL_DEVICE_STATUS = 'Device Status'

# ============================================================================

def normalize_mac_for_comparison(mac_address):
    """Normalize MAC address ONLY for comparison purposes - removes separators and uppercases"""
    if pd.isna(mac_address) or mac_address == '':
        return None
    
    # Convert to string and uppercase
    mac = str(mac_address).upper().strip()
    
    # Remove any separators for comparison only
    mac_clean = mac.replace(':', '').replace('-', '').replace('.', '').replace(' ', '')
    
    # Pad with zeros if less than 12 characters
    if len(mac_clean) < 12:
        mac_clean = mac_clean.zfill(12)
    
    # Take only first 12 characters if longer
    mac_clean = mac_clean[:12]
    
    # Validate it's hexadecimal
    try:
        int(mac_clean, 16)
    except ValueError:
        return None
    
    return mac_clean

# Define highlight colors
YELLOW_FILL = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')  # No name found
ORANGE_FILL = PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid')  # No switch info found
LIGHTBLUE_FILL = PatternFill(start_color='ADD8E6', end_color='ADD8E6', fill_type='solid')  # Duplicate MAC (same camera, different names)
RED_FILL = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')  # Duplicate IP with different MAC

def main():
    # Read the JSON file with camera inventory (switch info)
    print(f"Reading {CAMERA_INVENTORY_JSON}...")
    with open(CAMERA_INVENTORY_JSON, 'r') as f:
        camera_inventory_data = json.load(f)
    
    # Debug: Print the type and keys to understand the structure
    print(f"Data type: {type(camera_inventory_data)}")
    if isinstance(camera_inventory_data, dict):
        print(f"Keys in JSON: {list(camera_inventory_data.keys())}")
    
    # Handle both old format (list) and new format (dict with metadata)
    if isinstance(camera_inventory_data, dict):
        if 'cameras' in camera_inventory_data:
            # New format with metadata
            camera_inventory = camera_inventory_data['cameras']
            metadata = camera_inventory_data.get('discovery_metadata', {})
            
            print("\n" + "="*60)
            print("CAMERA DISCOVERY METADATA")
            print("="*60)
            print(f"Discovery timestamp: {metadata.get('timestamp', 'N/A')}")
            print(f"Seed switch: {metadata.get('seed_switch', 'N/A')}")
            print(f"Total cameras found: {metadata.get('total_cameras', len(camera_inventory))}")
            print(f"Total aggregates scanned: {metadata.get('total_aggregates', 'N/A')}")
            print("="*60 + "\n")
            
            # Print discovery statistics if available
            if 'discovery_statistics' in camera_inventory_data:
                stats = camera_inventory_data['discovery_statistics']
                print("DISCOVERY STATISTICS")
                print("="*60)
                print(f"Switches attempted: {stats.get('switches_attempted', 'N/A')}")
                print(f"Switches successfully scanned: {stats.get('switches_successfully_scanned', 'N/A')}")
                print(f"Switches failed: {stats.get('switches_failed_other', 0)}")
                
                if 'switches_by_type' in stats:
                    print("\nSwitches by type:")
                    for switch_type, type_stats in stats['switches_by_type'].items():
                        print(f"  {switch_type}:")
                        print(f"    - Attempted: {type_stats.get('attempted', 0)}")
                        print(f"    - Successful: {type_stats.get('successful', 0)}")
                        print(f"    - Failed: {type_stats.get('failed', 0)}")
                
                if 'failure_details' in stats and stats['failure_details']:
                    print("\nFailed switches:")
                    for failure in stats['failure_details']:
                        print(f"  - {failure.get('switch_name', 'Unknown')}: {failure.get('reason', 'Unknown reason')}")
                
                print("="*60 + "\n")
        else:
            # Dict but no 'cameras' key - might be old format stored as dict
            first_key = list(camera_inventory_data.keys())[0] if camera_inventory_data else None
            if first_key and isinstance(camera_inventory_data[first_key], dict):
                camera_inventory = list(camera_inventory_data.values())
            else:
                print("ERROR: Unexpected JSON structure")
                print(f"Please check the format of {CAMERA_INVENTORY_JSON}")
                return
    elif isinstance(camera_inventory_data, list):
        # Old format - just a list
        camera_inventory = camera_inventory_data
    else:
        print(f"ERROR: Unexpected data type in JSON: {type(camera_inventory_data)}")
        return
    
    # Create a dictionary with MAC address as key (for comparison) but store original MAC
    inventory_dict = {}
    switch_type_stats = {}
    
    for entry in camera_inventory:
        # Keep original MAC address format from JSON
        original_mac = entry['mac_address']
        # Create comparison key (normalized)
        mac_key = normalize_mac_for_comparison(original_mac)
        
        if mac_key:
            switch_type = entry.get('switch_type', 'UNKNOWN')
            inventory_dict[mac_key] = {
                'switch_name': entry['switch_name'],
                'switch_type': switch_type,
                'port': entry['port'],
                'vlan': entry.get('vlan', ''),
                'original_mac': original_mac  # Store original format
            }
            # Track statistics by switch type
            switch_type_stats[switch_type] = switch_type_stats.get(switch_type, 0) + 1
    
    print(f"Loaded {len(inventory_dict)} camera records from inventory")
    
    if switch_type_stats:
        print("\nCameras by switch type:")
        for switch_type, count in sorted(switch_type_stats.items()):
            print(f"  - {switch_type}: {count}")
    
    # Read the directory report Excel file
    print(f"\nReading {DIRECTORY_REPORT_EXCEL}...")
    directory_df = pd.read_excel(DIRECTORY_REPORT_EXCEL, sheet_name=DIRECTORY_SHEET_NAME)
    
    print(f"Total rows in directory report: {len(directory_df)}")
    
    # Check actual column names in the Excel file
    print(f"\nActual columns in Excel:")
    for col in directory_df.columns:
        print(f"  - {repr(col)}")
    
    # IMPORTANT: We only filter out formatting/status rows, NOT duplicate cameras
    original_count = len(directory_df)
    
    # Remove rows where camera name is NaN or empty (these are formatting rows)
    directory_df = directory_df[directory_df[DIR_COL_CAMERA_NAME].notna()]
    directory_df = directory_df[directory_df[DIR_COL_CAMERA_NAME].astype(str).str.strip() != '']
    
    filtered_count = len(directory_df)
    removed_count = original_count - filtered_count
    
    print(f"Filtered out {removed_count} empty/status formatting rows")
    print(f"Processing {filtered_count} camera records (including duplicates)")
    
    # Create comparison key for MACs in directory (for matching only)
    directory_df['MAC_comparison_key'] = directory_df[DIR_COL_MAC_ADDRESS].apply(normalize_mac_for_comparison)
    
    # Check for duplicate MACs to inform user
    duplicate_macs = directory_df[directory_df['MAC_comparison_key'].notna()].duplicated(subset=['MAC_comparison_key'], keep=False).sum()
    if duplicate_macs > 0:
        print(f"Note: Found {duplicate_macs} rows with duplicate MAC addresses (will be highlighted in light blue)")
    
    # Check for duplicate IPs with different MACs
    ip_groups = directory_df[directory_df[DIR_COL_IP_ADDRESS].notna()].groupby(DIR_COL_IP_ADDRESS)['MAC_comparison_key'].nunique()
    duplicate_ips = (ip_groups > 1).sum()
    if duplicate_ips > 0:
        print(f"WARNING: Found {duplicate_ips} IP addresses shared by different MAC addresses (will be highlighted in RED)")
    
    # Load the tracker workbook
    print(f"\nLoading {TRACKER_EXCEL}...")
    wb = load_workbook(TRACKER_EXCEL)
    ws = wb[TRACKER_CAMERA_SHEET]
    
    # Clear existing data (except header)
    print("Clearing existing data in camera sheet...")
    ws.delete_rows(2, ws.max_row)
    
    # Prepare data to write
    print("\nMatching cameras and preparing data...")
    row_num = 2
    matched_count = 0
    no_switch_info_count = 0
    no_name_info_count = 0
    duplicate_mac_count = 0
    duplicate_ip_count = 0
    
    # Track which MACs from inventory we've used
    used_macs = set()
    
    # Track MACs we've already written to detect duplicates
    written_macs = {}  # mac_key -> list of (row_num, camera_name, ip_address)
    
    # Track IPs we've already written to detect IP duplicates with different MACs
    written_ips = {}  # ip -> list of (row_num, mac_key)
    
    # First pass: Process all cameras from directory report
    for idx, cam_row in directory_df.iterrows():
        mac_key = cam_row['MAC_comparison_key']
        camera_name = cam_row[DIR_COL_CAMERA_NAME]
        ip_address = cam_row[DIR_COL_IP_ADDRESS]
        original_mac_from_dir = cam_row[DIR_COL_MAC_ADDRESS]
        
        # Check if this MAC was already written (duplicate detection)
        is_duplicate_mac = mac_key in written_macs if mac_key else False
        
        # Check if this MAC has a DIFFERENT IP than previous entries
        is_same_mac_diff_ip = False
        if is_duplicate_mac and ip_address:
            for prev_row, prev_name, prev_ip in written_macs[mac_key]:
                if prev_ip and prev_ip != ip_address:
                    is_same_mac_diff_ip = True
                    break
        
        # Check if this IP was already written with a DIFFERENT MAC
        is_duplicate_ip_diff_mac = False
        if ip_address and ip_address in written_ips:
            for prev_row, prev_mac in written_ips[ip_address]:
                if prev_mac != mac_key:
                    is_duplicate_ip_diff_mac = True
                    break
        
        # Look up switch info from inventory using comparison key
        switch_info = inventory_dict.get(mac_key)
        
        if switch_info:
            # Write to Excel - fully matched
            ws.cell(row=row_num, column=1, value=camera_name)
            # IMPORTANT: Force MAC to be text by setting number_format
            mac_cell = ws.cell(row=row_num, column=2, value=str(original_mac_from_dir))
            mac_cell.number_format = '@'  # '@' means text format in Excel
            ws.cell(row=row_num, column=3, value=str(ip_address))
            
            # Enhanced: Include switch type in the switch name display
            switch_type = switch_info.get('switch_type', 'UNKNOWN')
            switch_display = f"{switch_info['switch_name']} [{switch_type}]"
            ws.cell(row=row_num, column=4, value=switch_display)
            ws.cell(row=row_num, column=5, value=switch_info['port'])
            
            # Priority order for highlighting: RED > LIGHT BLUE
            if is_duplicate_ip_diff_mac or is_same_mac_diff_ip:
                # Network conflict - RED (most critical)
                for col in range(1, 6):
                    ws.cell(row=row_num, column=col).fill = RED_FILL
                duplicate_ip_count += 1
                
                if is_duplicate_ip_diff_mac:
                    for prev_row, prev_mac in written_ips[ip_address]:
                        if prev_mac != mac_key:
                            for col in range(1, 6):
                                ws.cell(row=prev_row, column=col).fill = RED_FILL
                    
                    print(f"  RED: Duplicate IP {ip_address} with different MAC:")
                    print(f"    - Current MAC: {original_mac_from_dir}")
                    print(f"    - Camera: {camera_name}")
                
                if is_same_mac_diff_ip:
                    for prev_row, prev_name, prev_ip in written_macs[mac_key]:
                        if prev_ip and prev_ip != ip_address:
                            for col in range(1, 6):
                                ws.cell(row=prev_row, column=col).fill = RED_FILL
                    
                    print(f"  RED: Same MAC {original_mac_from_dir} with different IP:")
                    print(f"    - Current IP: {ip_address}")
                    print(f"    - Camera: {camera_name}")
                
            elif is_duplicate_mac:
                # Duplicate MAC with same IP (same camera, different names) - LIGHT BLUE
                for col in range(1, 6):
                    ws.cell(row=row_num, column=col).fill = LIGHTBLUE_FILL
                duplicate_mac_count += 1
                prev_entries = written_macs[mac_key]
                print(f"  LIGHT BLUE: Duplicate MAC {original_mac_from_dir} found:")
                print(f"    - Previous: {prev_entries[-1][1]}")
                print(f"    - Current:  {camera_name}")
                # Also highlight the previous entries
                for prev_row, prev_name, prev_ip in prev_entries:
                    for col in range(1, 6):
                        ws.cell(row=prev_row, column=col).fill = LIGHTBLUE_FILL
            
            matched_count += 1
            if mac_key:
                used_macs.add(mac_key)
                if mac_key not in written_macs:
                    written_macs[mac_key] = []
                written_macs[mac_key].append((row_num, camera_name, ip_address))
            
            # Track IPs
            if ip_address:
                if ip_address not in written_ips:
                    written_ips[ip_address] = []
                written_ips[ip_address].append((row_num, mac_key))
                
        else:
            # Write partial data - have name but no switch info (highlight ORANGE)
            ws.cell(row=row_num, column=1, value=camera_name)
            # IMPORTANT: Force MAC to be text
            mac_cell = ws.cell(row=row_num, column=2, value=str(original_mac_from_dir))
            mac_cell.number_format = '@'
            ws.cell(row=row_num, column=3, value=str(ip_address))
            ws.cell(row=row_num, column=4, value='NOT FOUND')
            ws.cell(row=row_num, column=5, value='NOT FOUND')
            
            # Check for network conflicts even without switch info
            if is_duplicate_ip_diff_mac or is_same_mac_diff_ip:
                # RED takes priority over ORANGE
                for col in range(1, 6):
                    ws.cell(row=row_num, column=col).fill = RED_FILL
                duplicate_ip_count += 1
                
                if is_duplicate_ip_diff_mac:
                    for prev_row, prev_mac in written_ips[ip_address]:
                        if prev_mac != mac_key:
                            for col in range(1, 6):
                                ws.cell(row=prev_row, column=col).fill = RED_FILL
                    
                    print(f"  RED: Duplicate IP {ip_address} with different MAC (no switch info):")
                    print(f"    - Current MAC: {original_mac_from_dir}")
                
                if is_same_mac_diff_ip:
                    for prev_row, prev_name, prev_ip in written_macs[mac_key]:
                        if prev_ip and prev_ip != ip_address:
                            for col in range(1, 6):
                                ws.cell(row=prev_row, column=col).fill = RED_FILL
                    
                    print(f"  RED: Same MAC {original_mac_from_dir} with different IP (no switch info):")
                    print(f"    - Current IP: {ip_address}")
            else:
                # Highlight the entire row in ORANGE
                for col in range(1, 6):
                    ws.cell(row=row_num, column=col).fill = ORANGE_FILL
            
            no_switch_info_count += 1
            print(f"  ORANGE: No switch info found for {camera_name} (MAC: {original_mac_from_dir})")
            
            # Track written MACs and IPs even if no switch info found
            if mac_key:
                if mac_key not in written_macs:
                    written_macs[mac_key] = []
                written_macs[mac_key].append((row_num, camera_name, ip_address))
            
            if ip_address:
                if ip_address not in written_ips:
                    written_ips[ip_address] = []
                written_ips[ip_address].append((row_num, mac_key))
        
        row_num += 1
    
    # Second pass: Add cameras from inventory that don't have names (not in directory report)
    print("\nChecking for cameras in inventory without names...")
    for mac_key, switch_info in inventory_dict.items():
        if mac_key not in used_macs:
            # This MAC has switch info but no camera name (highlight YELLOW)
            ws.cell(row=row_num, column=1, value='NAME NOT FOUND')
            # IMPORTANT: Force MAC to be text - use the original MAC format from JSON
            mac_cell = ws.cell(row=row_num, column=2, value=switch_info['original_mac'])
            mac_cell.number_format = '@'
            ws.cell(row=row_num, column=3, value='')
            
            # Include switch type in display
            switch_type = switch_info.get('switch_type', 'UNKNOWN')
            switch_display = f"{switch_info['switch_name']} [{switch_type}]"
            ws.cell(row=row_num, column=4, value=switch_display)
            ws.cell(row=row_num, column=5, value=switch_info['port'])
            
            # Highlight the entire row in YELLOW
            for col in range(1, 6):
                ws.cell(row=row_num, column=col).fill = YELLOW_FILL
            
            no_name_info_count += 1
            print(f"  YELLOW: No camera name found for MAC: {switch_info['original_mac']} on {switch_info['switch_name']} [{switch_type}] port {switch_info['port']}")
            
            row_num += 1
    
    # Save the workbook
    print(f"\nSaving updated tracker to {OUTPUT_EXCEL}...")
    wb.save(OUTPUT_EXCEL)
    
    # Print summary
    print("\n" + "="*60)
    print("SUMMARY")
    print("="*60)
    print(f"Total cameras from directory: {len(directory_df)}")
    print(f"Total MACs from inventory: {len(inventory_dict)}")
    print(f"\n✓ Successfully matched: {matched_count}")
    print(f"⚠  ORANGE - Have name but no switch info: {no_switch_info_count}")
    print(f"⚠  YELLOW - Have switch info but no name: {no_name_info_count}")
    print(f"⚠  LIGHT BLUE - Duplicate MAC (same camera, different names): {duplicate_mac_count}")
    print(f"⚠  RED - Duplicate IP with different MACs (IP CONFLICT): {duplicate_ip_count}")
    print(f"\nTotal rows written: {row_num - 2}")
    print(f"\nOutput saved to: {OUTPUT_EXCEL}")
    print("\nColor coding:")
    print("  - No color: Fully matched (name + switch info)")
    print("  - ORANGE: Camera name found but no switch info")
    print("  - YELLOW: Switch info found but no camera name")
    print("  - LIGHT BLUE: Same MAC/IP with different names (needs reconciliation)")
    print("  - RED: Network conflict (CRITICAL!)")
    print("      * Same IP with different MACs, OR")
    print("      * Same MAC with different IPs")
    print("\nNOTE: Switch names now include switch type in brackets [EDGE/SERVER/OTHER]")
    print("      Red entries indicate network configuration errors that need immediate attention!")
    print("="*60)

if __name__ == "__main__":
    main()
